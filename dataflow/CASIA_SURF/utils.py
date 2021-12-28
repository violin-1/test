import os
import torch
import numpy as np
import pandas as pd
from openpyxl import load_workbook
import torch.optim
import torch.utils.data
import torch.optim
from torch.autograd import Variable
from dataflow.log import logger
from tqdm import tqdm

def one_hot(batchsize,numclass,tensor):
    return torch.zeros(batchsize,numclass).to(tensor.device).scatter_(1,tensor.view(1,tensor.shape[0]).long().t(),1)

def train_init(Trainer,datactr):
    Trainer.log = logger(datactr.config.arch,datactr.config.datatsetstype[datactr.curdataset],datactr.datasets[datactr.curdataset].out,datactr.config.trainratio)
    Trainer.log.log1()
    Trainer.savedir = Trainer.log.savedir
    Trainer.best_res = None
    Trainer.set(datactr)
    
def train_epoch(Trainer):

    Trainer.model.train()
    Trainer.dataset.set(phase_train = True)
    Trainer.dataloader_train = torch.utils.data.DataLoader(Trainer.dataset, batch_size=Trainer.batch_size, shuffle=True,num_workers=Trainer.num_workers)
    
    Trainer.log.log2(Trainer.epoch)
    acc = 0
    for i, (s0, s1, label) in enumerate(tqdm(Trainer.dataloader_train)):
        # measure data loading time
        label = one_hot(s0.shape[0],2,label)
        s0 = Variable(s0).float().to(Trainer.device)
        s1 = Variable(s1).float().to(Trainer.device)
        label = Variable(label).float().to(Trainer.device)
        
        # compute output
        pred = Trainer.model(s0, s1)
        Trainer.loss = Trainer.criterion(pred,label)
        
        Trainer.loss.backward()
        Trainer.optimizer.step()
        acc += torch.where(Trainer.resdeal(pred)==Trainer.resdeal(label))[0].shape[0]
        
    train_acc = acc/Trainer.dataset.__len__()
    Trainer.log.log3(Trainer.epoch,train_acc,Trainer.loss)

def val_epoch(Trainer):
    TP,TN,FP,FN = {},{},{},{}
    Trainer.model.eval()
    with torch.no_grad():
        ###########################################
        '''                val             '''
        ###########################################
        # val for threshold
        Trainer.dataset.set(phase_train = False)
        Trainer.dataloader_val = torch.utils.data.DataLoader(Trainer.dataset, batch_size=Trainer.batch_size, shuffle=True,num_workers=Trainer.num_workers)
        
        TP['all'],TN['all'],FP['all'],FN['all']  = 0,0,0,0
        for key in Trainer.dataset.cat_name.keys():
            TP[key],TN[key],FP[key],FN[key]  = 0,0,0,0
        for i, (s0, s1, label, cat) in enumerate(tqdm(Trainer.dataloader_val)):
            # measure data loading time
            # label = one_hot(s0.shape[0],2,label)
            s0 = Variable(s0).float().to(Trainer.device)
            s1 = Variable(s1).float().to(Trainer.device)
            
            Trainer.optimizer.zero_grad()
            # compute output
            pred = Trainer.model(s0, s1)
            pred = Trainer.resdeal(pred)
            
            label = Variable(label).float().to(Trainer.device)
            cat = Variable(cat).float().to(Trainer.device)
            
            TP['all'] += torch.where((pred==1) &(label == 1))[0].shape[0]
            TN['all'] += torch.where((pred==0) &(label == 0))[0].shape[0]
            FP['all'] += torch.where((pred==1) &(label == 0))[0].shape[0]
            FN['all'] += torch.where((pred==0) &(label == 1))[0].shape[0]
            
            for key in Trainer.dataset.cat_name.keys():
                TP[key] += torch.where((cat == Trainer.dataset.cat_name[key])&(pred==1) &(label == 1))[0].shape[0]
                TN[key] += torch.where((cat == Trainer.dataset.cat_name[key])&(pred==0) &(label == 0))[0].shape[0]
                FP[key] += torch.where((cat == Trainer.dataset.cat_name[key])&(pred==1) &(label == 0))[0].shape[0]
                FN[key] += torch.where((cat == Trainer.dataset.cat_name[key])&(pred==0) &(label == 1))[0].shape[0]
    # if (TP['all']+TN['all']+FP['all']+FN['all']) == 0:
    #     Trainer.acc=0
    # else:
    #     Trainer.acc = (TP['all'] + TN['all'])/(TP['all']+TN['all']+FP['all']+FN['all'])
    csvfile = os.path.join('outputs',Trainer.log.modeltype+'_'+Trainer.log.datasettype+'_'+ Trainer.trainratio+'.xlsx')
    # Trainer.args_val.append(print(TP,TN,FP,FN,Trainer.dataset.cat_name,Trainer.dataset.out,csvfile))
    Trainer.args_val = get_res(TP,TN,FP,FN,Trainer.dataset.cat_name,Trainer.dataset.out,csvfile)
    Trainer.log.log4(Trainer.epoch,Trainer.args_val)

def get_res(TP,TN,FP,FN,Tname,outname,csvfile):
    Accuracy,Precision,Recall,F1_score,APCER,NPCER,ACER = {},{},{},{},{},{},{}
    ans = np.zeros((1,7))
    # TP['all'],TN['all'],FP['all'],FN['all']  = 0,0,0,0
    for i,key in enumerate(TP.keys()):
        if (TP[key]+TN[key]+FP[key]+FN[key]) == 0:
            Accuracy[key]=0
        else:
            Accuracy[key] = (TP[key] + TN[key])/(TP[key]+TN[key]+FP[key]+FN[key])
        if (TP[key]+FP[key]) == 0:
            Precision[key]=0
        else:
            if key == 'all' or key == '真人':
                Precision[key] = TP[key]/(TP[key]+FP[key])
            else:
                Precision[key] = TN[key]/(TN[key]+FN[key])
        if (TP[key]+FN[key]) == 0:
            Recall[key]=0
        else:
            if key == 'all' or key == '真人':
                Recall[key] = TP[key]/(TP[key]+FN[key])
            else:
                Recall[key] = TN[key]/(TN[key]+FP[key])
        if (Precision[key] + Recall[key]) == 0:
            F1_score[key]=0
        else:
            F1_score[key] = 2 * Precision[key] * Recall[key] / (Precision[key] + Recall[key])
        if (TN[key]+FP[key]) == 0:
            APCER[key]=0
        else:
            APCER[key] = FP[key]/(TN[key]+FP[key])
        if (FN[key]+TP[key]) == 0:
            NPCER[key]=0
        else:
            NPCER[key] = FN[key]/(FN[key]+TP[key])
        
        ACER[key] = (APCER[key] + NPCER[key])/2
        ans[i,0],ans[i,1],ans[i,2],ans[i,3],ans[i,4],ans[i,5],ans[i,6], = Accuracy[key],Precision[key],Recall[key],F1_score[key],APCER[key],NPCER[key],ACER[key]
    
    data = pd.DataFrame(ans.T)
    data.columns = ['all']+list(Tname)
    args_name = ['Accurate','Precision','Recall','F1 Score','APCER','NPCER','ACER']
    
    data.insert(0,outname,args_name)

    book = None
    if os.path.exists(csvfile):
        book = load_workbook(csvfile)
    writer = pd.ExcelWriter(csvfile)
        # writer.book = load_workbook(os.path.join('outputs',log.modeltype+'_'+log.datasettype+'_'+ datactr.config.trainratio+'.xlsx'))
    if book is not None:
        writer.book = book
    data.to_excel(excel_writer=writer,sheet_name=outname,float_format='%.6f',index=False)
    writer.save()
    writer.close()
    
    args_dict = {}
    args_dict['Accurate'] = Accuracy
    args_dict['Precision'] = Precision
    args_dict['Recall'] = Recall
    args_dict['F1_score'] = F1_score
    args_dict['APCER'] = APCER
    args_dict['NPCER'] = NPCER
    args_dict['ACER'] = ACER

    return args_dict
    # print(Accuracy)
    # print(Precision)
    # print(Recall)
    # print(F1_score)
    # print(APCER)
    # print(NPCER)
    # print(ACER)

def res_compare(Trainer):
    Trainer.is_best = False
    if Trainer.best_res is None or Trainer.args_val['Accurate']['all'] > Trainer.best_res['Accurate']['all']:
        Trainer.best_res = Trainer.args_val
        Trainer.is_best = True
    Trainer.log.log_best(Trainer.epoch,Trainer.args_val,Trainer.best_res)

def end_func(Trainer):
    Trainer.log.log5()