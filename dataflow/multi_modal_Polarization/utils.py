import torch
from torch.utils import data
import os
import numpy as np
from tqdm import tqdm
import pandas as pd
from openpyxl import load_workbook
from torch.autograd.variable import Variable
from torch.utils.data import DataLoader
from dataflow.log import logger

def one_hot(batchsize,numclass,tensor):
    return torch.zeros(batchsize,numclass).to(tensor.device).scatter_(1,tensor.view(1,tensor.shape[0]).long().t(),1)

def drawGraph(datactr):
    logfile = datactr.config.draw_logfile
    logpath = os.path.join('model',datactr.config.arch,'logs',logfile)
    best_acc = 0
    valacc = []
    trainloss = []
    epoch = []
    def abstr(line, strfind, strend): 
        return line[line.find(strfind)+len(strfind):line.find(strfind)+len(strfind)+line[line.find(strfind)+len(strfind):].find(strend)]
    
    with open(logpath,'r') as f:
        for line in f.readlines():
            if '||epoch:' in line:
                tmp = abstr(line, '||epoch:', ',')
                epoch.append(int(tmp))
                tmp = abstr(line,'Val_accurate = ',',')
                if(float(tmp) > best_acc):
                    best_acc = float(tmp)
                valacc.append(best_acc)
                tmp = abstr(line,'Train_loss = ',' ')
                trainloss.append(float(tmp))
    valacc = np.array(valacc)*100
    import matplotlib.pyplot as plt
    plt.figure(figsize=[10,5])
    plt.plot(epoch, valacc, label='CDCN_MultiModality')
    plt.title('Val_Acc_Curve best_acc:{}'.format(best_acc))
    plt.xlabel('epoch')
    plt.ylabel('Accurate(%)')
    plt.savefig(os.path.join('outputs','CDCN_MultiModality_Val_Acc.png'))
    
    plt.figure(figsize=[10,5])
    plt.plot(epoch, trainloss, label='CDCN_MultiModality')
    plt.title('Train_Loss_Curve')
    plt.xlabel('epoch')
    plt.ylabel('Loss')
    plt.savefig(os.path.join('outputs','CDCN_MultiModality_Train_Loss.png'))
    pass

def train_init(Trainer,datactr):
    Trainer.log = logger(datactr.config.arch,datactr.config.datatsetstype[datactr.curdataset],datactr.datasets[datactr.curdataset].out,datactr.config.trainratio)
    Trainer.log.log1()
    Trainer.savedir = Trainer.log.savedir
    Trainer.best_res = None
    Trainer.set(datactr)
    
def multi_train_epoch(Trainer):
    Trainer.model.train()
    Trainer.dataset.set(dtype = 'train')
    Trainer.dataloader_train = DataLoader(Trainer.dataset,batch_size= Trainer.batch_size, shuffle=True,num_workers=Trainer.num_workers)
    
    Trainer.log.log2(Trainer.epoch)
    acc = 0
    for i,(s0,dolp,label) in enumerate(tqdm(Trainer.dataloader_train)):
        label = one_hot(s0.shape[0],2,label)
        s0,dolp,label = Variable(s0).float().to(Trainer.device),Variable(dolp).float().to(Trainer.device),Variable(label).float().to(Trainer.device)
        Trainer.optimizer.zero_grad()
        
        pred =  Trainer.model(s0, dolp)

        Trainer.loss = Trainer.criterion(pred,label)
        
        Trainer.loss.backward()
    
        Trainer.optimizer.step()
        acc += torch.where(Trainer.resdeal(pred)==Trainer.resdeal(label))[0].shape[0]
    
    train_acc = acc/Trainer.dataset.__len__()
    Trainer.log.log3(Trainer.epoch,train_acc,Trainer.loss)
    pass

def multi_val_epoch(Trainer):
    TP,TN,FP,FN = {},{},{},{}
    Trainer.model.eval()
    with torch.no_grad():
        ###########################################
        '''                val             '''
        ###########################################
        # val for threshold
        Trainer.dataset.set(dtype = 'test')
        Trainer.dataloader_val = DataLoader(Trainer.dataset,batch_size=Trainer.batch_size, shuffle=True,num_workers=Trainer.num_workers)
        
        TP['all'],TN['all'],FP['all'],FN['all']  = 0,0,0,0
        for key in Trainer.dataset.cat_name.keys():
            TP[key],TN[key],FP[key],FN[key]  = 0,0,0,0
        for i,(s0,dolp,label,cat) in enumerate(tqdm(Trainer.dataloader_val)):
            # get the inputs
            s0,dolp = Variable(s0).float().to(Trainer.device),Variable(dolp).float().to(Trainer.device)
            Trainer.optimizer.zero_grad()
            
            pred = Trainer.model(s0, dolp)
            pred = Trainer.resdeal(pred)
            
            cat = cat.to(Trainer.device)
            label = label.to(Trainer.device)
            TP['all'] += torch.where((pred==1) &(label == 1))[0].shape[0]
            TN['all'] += torch.where((pred==0) &(label == 0))[0].shape[0]
            FP['all'] += torch.where((pred==1) &(label == 0))[0].shape[0]
            FN['all'] += torch.where((pred==0) &(label == 1))[0].shape[0]
            for key in Trainer.dataset.cat_name.keys():
                TP[key] += torch.where((cat == Trainer.dataset.cat_name[key])&(pred==1) &(label == 1))[0].shape[0]
                TN[key] += torch.where((cat == Trainer.dataset.cat_name[key])&(pred==0) &(label == 0))[0].shape[0]
                FP[key] += torch.where((cat == Trainer.dataset.cat_name[key])&(pred==1) &(label == 0))[0].shape[0]
                FN[key] += torch.where((cat == Trainer.dataset.cat_name[key])&(pred==0) &(label == 1))[0].shape[0]
    
    csvfile = os.path.join('outputs',Trainer.log.modeltype+'_'+Trainer.log.datasettype+'_'+ Trainer.trainratio+'.xlsx')
    Trainer.args_val = get_res(TP,TN,FP,FN,Trainer.dataset.cat_name,Trainer.dataset.out,csvfile)
    Trainer.log.log4(Trainer.epoch,Trainer.args_val)

def single_train_epoch(Trainer):

    Trainer.model.train()
    Trainer.dataset.set(dtype = 'train')
    Trainer.dataloader_train = DataLoader(Trainer.dataset,batch_size= Trainer.batch_size, shuffle=True,num_workers=Trainer.num_workers)
    
    Trainer.log.log2(Trainer.epoch)
    acc = 0
    for i,(single,label) in enumerate(tqdm(Trainer.dataloader_train)):
        label = one_hot(single.shape[0],2,label)
        single,label = Variable(single).float().to(Trainer.device),Variable(label).float().to(Trainer.device)
        Trainer.optimizer.zero_grad()
        
        pred =  Trainer.model(single)

        Trainer.loss = Trainer.criterion(pred,label)
        
        Trainer.loss.backward()
    
        Trainer.optimizer.step()
        acc += torch.where(Trainer.resdeal(pred)==Trainer.resdeal(label))[0].shape[0]
    
    train_acc = acc/Trainer.dataset.__len__()
    Trainer.log.log3(Trainer.epoch,train_acc,Trainer.loss)
    
def single_val_epoch(Trainer):
    TP,TN,FP,FN = {},{},{},{}
    # Accuracy,Precision,Recall,F1_score,APCER,NPCER,ACER = {},{},{},{},{},{},{}
    Trainer.model.eval()
    with torch.no_grad():
        ###########################################
        '''                val             '''
        ###########################################
        # val for threshold
        Trainer.dataset.set(dtype = 'test')
        Trainer.dataloader_val = DataLoader(Trainer.dataset,batch_size=Trainer.batch_size, shuffle=True,num_workers=Trainer.num_workers)
        
        TP['all'],TN['all'],FP['all'],FN['all']  = 0,0,0,0
        for key in Trainer.dataset.cat_name.keys():
            TP[key],TN[key],FP[key],FN[key]  = 0,0,0,0
        for i,(single,label,cat) in enumerate(tqdm(Trainer.dataloader_val)):
            # get the inputs
            single = Variable(single).float().to(Trainer.device)
            Trainer.optimizer.zero_grad()
            
            pred = Trainer.model(single)
            pred = Trainer.resdeal(pred)

            cat = cat.to(Trainer.device)
            label = label.to(Trainer.device)
            TP['all'] += torch.where((pred==1) &(label == 1))[0].shape[0]
            TN['all'] += torch.where((pred==0) &(label == 0))[0].shape[0]
            FP['all'] += torch.where((pred==1) &(label == 0))[0].shape[0]
            FN['all'] += torch.where((pred==0) &(label == 1))[0].shape[0]
            for key in Trainer.dataset.cat_name.keys():
                TP[key] += torch.where((cat == Trainer.dataset.cat_name[key])&(pred==1) &(label == 1))[0].shape[0]
                TN[key] += torch.where((cat == Trainer.dataset.cat_name[key])&(pred==0) &(label == 0))[0].shape[0]
                FP[key] += torch.where((cat == Trainer.dataset.cat_name[key])&(pred==1) &(label == 0))[0].shape[0]
                FN[key] += torch.where((cat == Trainer.dataset.cat_name[key])&(pred==0) &(label == 1))[0].shape[0]

    csvfile = os.path.join('outputs',Trainer.log.modeltype+'_'+Trainer.log.datasettype+'_'+ Trainer.trainratio+'.xlsx')
    Trainer.args_val = get_res(TP,TN,FP,FN,Trainer.dataset.cat_name,Trainer.dataset.out,csvfile)
    Trainer.log.log4(Trainer.epoch,Trainer.args_val)

def get_res(TP,TN,FP,FN,Tname,outname,csvfile):
    Accuracy,Precision,Recall,F1_score,APCER,NPCER,ACER = {},{},{},{},{},{},{}
    ans = np.zeros((7,7))
    # TP['all'],TN['all'],FP['all'],FN['all']  = 0,0,0,0
    for i,key in enumerate(TP.keys()):
        if (TP[key]+TN[key]+FP[key]+FN[key]) == 0:
            Accuracy[key]=0
        else:
            Accuracy[key] = (TP[key] + TN[key])/(TP[key]+TN[key]+FP[key]+FN[key])
            
        if key == 'all' or key == '真人':
            if (TP[key]+FP[key]) == 0:
                Precision[key]=0
            else:
                Precision[key] = TP[key]/(TP[key]+FP[key])
        else:
            if (TN[key]+FN[key]) == 0:
                Precision[key]=0
            else:
                Precision[key] = TN[key]/(TN[key]+FN[key])
                
        if key == 'all' or key == '真人':
            if (TP[key]+FN[key]) == 0:
                Recall[key]=0
            else:
                Recall[key] = TP[key]/(TP[key]+FN[key])
        else:
            if (TN[key]+FP[key]) == 0:
                Recall[key]=0
            else:
                Recall[key] = TN[key]/(TN[key]+FP[key])
                
        if (Precision[key] + Recall[key]) == 0:
            F1_score[key]=0
        else:
            F1_score[key] = 2 * Precision[key] * Recall[key] / (Precision[key] + Recall[key])
            
        if (TN[key]+FP[key]) == 0:
            APCER[key]=0
        else:
            APCER[key] = FP[key]/(TN[key]+FP[key])
        
        if (FN[key]+TP[key]) == 0:
            NPCER[key]=0
        else:
            NPCER[key] = FN[key]/(FN[key]+TP[key])
        
        ACER[key] = (APCER[key] + NPCER[key])/2
        ans[i,0],ans[i,1],ans[i,2],ans[i,3],ans[i,4],ans[i,5],ans[i,6], = Accuracy[key],Precision[key],Recall[key],F1_score[key],APCER[key],NPCER[key],ACER[key]
    
    data = pd.DataFrame(ans.T)
    data.columns = ['all']+list(Tname)
    args_name = ['Accurate','Precision','Recall','F1 Score','APCER','NPCER','ACER']
    data.insert(0,outname,args_name)

    book = None
    if os.path.exists(csvfile):
        book = load_workbook(csvfile)
    writer = pd.ExcelWriter(csvfile)
        # writer.book = load_workbook(os.path.join('outputs',log.modeltype+'_'+log.datasettype+'_'+ datactr.config.trainratio+'.xlsx'))
    if book is not None:
        writer.book = book
    data.to_excel(excel_writer=writer,sheet_name=outname,float_format='%.6f',index=False)
    writer.save()
    writer.close()
    
    args_dict = {}
    args_dict['Accurate'] = Accuracy
    args_dict['Precision'] = Precision
    args_dict['Recall'] = Recall
    args_dict['F1_score'] = F1_score
    args_dict['APCER'] = APCER
    args_dict['NPCER'] = NPCER
    args_dict['ACER'] = ACER
    
    return args_dict
    # print(Accuracy)
    # print(Precision)
    # print(Recall)
    # print(F1_score)
    # print(APCER)
    # print(NPCER)
    # print(ACER)
    
def res_compare(Trainer):
    Trainer.is_best = False
    if Trainer.best_res  is None or Trainer.args_val['Accuracy']['all'] > Trainer.best_res['Accuracy']['all']:
        Trainer.best_res = Trainer.args_val
        Trainer.is_best = True
    Trainer.log.log_best(Trainer.epoch,Trainer.args_val,Trainer.best_res)

def end_func(Trainer):
    Trainer.log.log5()